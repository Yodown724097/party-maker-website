"""
sync_ramadan_products.py
把 PM PRODUCTS_产品总表_All Ramadan.xlsx 转换成模板格式，
上传新品图片到 R2（使用 upload_r2.py 的正确凭证），合并到 products.json
"""

import json
import os
import sys
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import boto3
from botocore.config import Config

# ============ 配置（从 upload_r2.py 复制的正确凭证）============
EXCEL_IN   = r'C:\Users\Administrator\Downloads\PM PRODUCTS_产品总表_All Ramadan.xlsx'
TEMPLATE    = r'c:\Users\Administrator\WorkBuddy\20260423171832\产品数据上传模板_v2.xlsx'
EXCEL_OUT   = r'c:\Users\Administrator\WorkBuddy\20260423171832\产品数据上传模板_v2.xlsx'
PRODUCTS_JSON = r'c:\Users\Administrator\WorkBuddy\20260423171832\party-maker-website\products.json'
UPLOAD_DIR   = r'D:\upload'

R2_ENDPOINT = 'https://cdd100719805df54e62bee48d165b2dd.r2.cloudflarestorage.com'
R2_ACCESS_KEY = '6ba9614989d68d1b8f7f7d6b53f50e54'
R2_SECRET_KEY = '10d4b41750b6965866db2bac4f33c8d6be56679219efe4cab6ae0211eacd6d80'
R2_BUCKET  = 'party-maker'
R2_PUBLIC_BASE = 'https://pub-1fd965ab66464286847edcb540254451.r2.dev'

COLMAP = {
    'Item No.':     'ID/SKU',
    'Product Name':  'Product Name',
    'USD Price':    'Price (USD)',
    'Theme':        'Theme',
    'Function':     'Subcategory',
    'Description':  'Description',
    '成本价':       'Cost Price (CNY)',
    '成本备注':     'Cost Note',
    '下达单号':     'Order No',
    'Stock Qty':   'Stock Qty',
    'Unit Size':   'Unit Size',
    'CTN L':       'CTN Length (cm)',
    'CTN W':       'CTN Width (cm)',
    'CTN H':       'CTN Height (cm)',
    'pcs/CTN':     'pcs/CTN',
    'CBM':         'CBM',
    'N.W':         'N.W (kg)',
    'G.W':         'G.W (kg)',
    'Tags':        'Tags',
}

TEMPLATE_COLS = [
    'ID/SKU','Product Name','Price (USD)','Theme','Subcategory',
    'Description','Images URLs','Cost Price (CNY)','Cost Note',
    'Order No','Stock Qty','Unit Size',
    'CTN Length (cm)','CTN Width (cm)','CTN Height (cm)',
    'pcs/CTN','CBM','N.W (kg)','G.W (kg)','Tags'
]

def get_r2_client():
    return boto3.client(
        's3',
        endpoint_url=R2_ENDPOINT,
        aws_access_key_id=R2_ACCESS_KEY,
        aws_secret_access_key=R2_SECRET_KEY,
        config=Config(s3={'use_accelerate': False}),
    )

def upload_one(s3, sku, fname):
    local = os.path.join(UPLOAD_DIR, fname)
    if not os.path.exists(local):
        print('    [warn] 本地不存在: ' + fname)
        return None
    key = sku + '/' + fname
    try:
        s3.upload_file(local, R2_BUCKET, key,
                       ExtraArgs={'ContentType': 'image/jpeg', 'CacheControl': 'public, max-age=31536000'})
        return R2_PUBLIC_BASE + '/' + key
    except Exception as e:
        print('    [error] 上传失败 ' + fname + ': ' + str(e))
        return None

def main():
    print('=' * 60)
    print('Sync Ramadan Products -> products.json')
    print('=' * 60)

    # 1. 读取现有 products.json
    print('\n[1/5] 读取现有 products.json...')
    with open(PRODUCTS_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    existing = data if isinstance(data, list) else data.get('products', [])
    exist_dict = {p['sku']: p for p in existing}
    print('  现有产品数: ' + str(len(exist_dict)))

    # 2. 读取用户 Excel
    print('\n[2/5] 读取 Excel: ' + os.path.basename(EXCEL_IN))
    df = pd.read_excel(EXCEL_IN, sheet_name='产品总表')
    df.columns = [str(c).strip() for c in df.columns]
    print('  行数: ' + str(len(df)))

    # 3. 初始化 R2
    print('\n[3/5] 连接 R2...')
    s3 = get_r2_client()
    print('  R2 连接成功')

    # 4. 处理每一行
    print('\n[4/5] 处理产品数据...')
    new_count = 0
    update_count = 0
    rows_out = []

    for _, row in df.iterrows():
        sku = str(row.get('Item No.', '')).strip()
        if not sku or sku == 'nan':
            continue

        is_new = sku not in exist_dict
        if is_new:
            new_count += 1
            print('  [new] ' + sku)
        else:
            update_count += 1

        out = {}
        for src, dst in COLMAP.items():
            val = row.get(src)
            if pd.isna(val) or str(val).strip() in ('', 'nan'):
                out[dst] = ''
            else:
                out[dst] = str(val).strip()

        # 处理图片
        img_str = row.get('Image', '')
        urls = []
        if not pd.isna(img_str) and str(img_str).strip():
            fnames = [f.strip() for f in str(img_str).split(',') if f.strip()]
            for fname in fnames:
                local_path = os.path.join(UPLOAD_DIR, fname)
                if os.path.exists(local_path):
                    url = upload_one(s3, sku, fname)
                    if url:
                        urls.append(url)
                else:
                    urls.append(R2_PUBLIC_BASE + '/' + sku + '/' + fname)
        if not urls and not is_new and sku in exist_dict:
            urls = exist_dict[sku].get('images', [])
        out['Images URLs'] = ';'.join(urls)

        for c in TEMPLATE_COLS:
            if c not in out:
                out[c] = ''
        rows_out.append(out)

    print('  新品: ' + str(new_count) + ', 更新: ' + str(update_count))

    # 5. 保存到模板 Excel
    print('\n[5/5] 保存到模板 Excel...')
    wb = load_workbook(TEMPLATE)
    if '产品数据' in wb.sheetnames:
        del wb['产品数据']
    ws = wb.create_sheet('产品数据', 0)
    ws.append(TEMPLATE_COLS)
    for r in rows_out:
        ws.append([r.get(c, '') for c in TEMPLATE_COLS])
    wb.save(EXCEL_OUT)
    print('  保存成功: ' + EXCEL_OUT)

    # 6. 合并到 products.json
    print('\n[6/6] 合并到 products.json...')
    merged = exist_dict.copy()
    for r in rows_out:
        sku = r['ID/SKU']
        prod = {}
        field_map = {
            'ID/SKU': 'sku',
            'Product Name': 'name',
            'Price (USD)': 'price',
            'Theme': 'theme',
            'Subcategory': 'subcategory',
            'Description': 'description',
            'Images URLs': 'images',
            'Tags': 'tags',
            'Cost Price (CNY)': '_costPrice',
            'Cost Note': '_costNote',
            'Order No': '_orderNo',
            'Stock Qty': '_stockQty',
            'Unit Size': '_unitSize',
            'CTN Length (cm)': '_ctnL',
            'CTN Width (cm)': '_ctnW',
            'CTN Height (cm)': '_ctnH',
            'pcs/CTN': '_pcsPerCtn',
            'CBM': '_cbm',
            'N.W (kg)': '_nw',
            'G.W (kg)': '_gw',
        }
        for col, field in field_map.items():
            val = r.get(col, '')
            if col == 'Images URLs':
                prod[field] = [u.strip() for u in str(val).split(';') if u.strip()] if val else []
            elif col == 'Tags' and val:
                prod[field] = [t.strip().lower() for t in str(val).split(',') if t.strip()]
            elif field in ('price', '_costPrice'):
                try:
                    prod[field] = float(val) if val else 0
                except:
                    prod[field] = 0
            elif field in ('_ctnL','_ctnW','_ctnH','_pcsPerCtn','_cbm','_nw','_gw','_stockQty'):
                try:
                    prod[field] = float(val) if val else 0
                except:
                    prod[field] = 0
            else:
                prod[field] = val
        merged[sku] = prod

    merged_list = list(merged.values())
    out_data = {
        'products': merged_list,
        'themes': sorted(set(p.get('theme','') for p in merged_list if p.get('theme'))),
        'categories': sorted(set(p.get('subcategory','') for p in merged_list if p.get('subcategory'))),
        'lastUpdated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    with open(PRODUCTS_JSON, 'w', encoding='utf-8') as f:
        json.dump(out_data, f, ensure_ascii=False, indent=2)
    print('  products.json 已更新 (' + str(len(merged_list)) + ' 个产品)')

    print('\nDone! 下一步: 运行 build_pages.py 重新生成页面')

if __name__ == '__main__':
    main()
