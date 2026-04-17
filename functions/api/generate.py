"""
Cloudflare Pages Function - PI Excel Generator
流程:
  1. 生成Excel（含Cost）
  2. 发客户通知邮件（无附件，无成本价）
  3. POST Excel到VPS → VPS发info邮件（含附件含成本价）
"""
import json
import base64
import io
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

VPS_URL = "https://api.partymaker.cn/trigger-mail"
TRIGGER_KEY = "pm-trigger-2026"

SMTP_SERVER = "smtp.qiye.aliyun.com"
SMTP_PORT = 465
SMTP_USER = "info@partymaker.cn"
SMTP_PASS = "JT.info1805"
FROM_EMAIL = "info@partymaker.cn"


def generate_pi_no():
    now = datetime.now()
    return f'PI-{str(now.year)[-2:]}{now.strftime("%m%d")}-{random.randint(1000, 9999)}'


def get_cell_letter(col_idx):
    result = ''
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def on_request(request, env):
    if request.method == 'OPTIONS':
        return cors_response()
    if request.method != 'POST':
        return json_response({'error': 'Method not allowed'}, status=405)

    try:
        body = request.json
    except:
        return json_response({'error': 'Invalid JSON'}, status=400)

    contact = body.get('contact', {})
    cart = body.get('cart', [])
    send_email_flag = body.get('send_email', True)

    if not contact.get('name') or not contact.get('email'):
        return json_response({'error': 'Missing name/email'}, status=400)
    if not cart:
        return json_response({'error': 'Cart is empty'}, status=400)

    now = datetime.now()
    pi_no = generate_pi_no()

    # ===== 生成Excel（带Cost） =====
    excel_buffer = generate_excel(contact, cart, pi_no, now)
    excel_bytes = excel_buffer.getvalue()
    excel_b64 = base64.b64encode(excel_bytes).decode('utf-8')

    results = {}

    # ===== 1. 发客户通知邮件（无附件，无成本价） =====
    if send_email_flag:
        customer_result = send_customer_email(contact, cart, pi_no, excel_bytes)
        results['customer_email'] = customer_result

        # ===== 2. POST给VPS → VPS发info邮件（含附件含成本价） =====
        vps_result = notify_vps({
            'pi_no': pi_no,
            'excel_b64': excel_b64,
            'contact': contact,
            'cart': cart,
            'total': sum(int(i.get('quantity', 0) or 0) * float(i.get('price', 0) or 0) for i in cart)
        })
        results['owner_email'] = vps_result

    return json_response({
        'success': True,
        'piNo': pi_no,
        'excel': excel_b64,
        'results': results,
    })


def send_customer_email(contact, cart, pi_no, excel_bytes):
    """发客户通知邮件——无附件，无成本价，纯通知"""
    import urllib.request
    import urllib.error
    import smtplib
    import ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    to_email = contact.get('email', '')
    if not to_email:
        return {'sent': False, 'error': 'No customer email'}

    # 构建产品列表（无Cost，无附件）
    rows_html = ""
    rows_text = ""
    total = 0
    for i, item in enumerate(cart, 1):
        qty = int(item.get('quantity', 0) or 0)
        price = float(item.get('price', 0) or 0)
        subtotal = qty * price
        total += subtotal
        sku = item.get('sku', item.get('id', '-'))
        name = item.get('name', '')
        img = ((item.get('images') or [''])[0] or '') if isinstance(item.get('images'), list) else ''
        unit_size = item.get('_unitSize', '')
        if str(unit_size) in ['nan', None, '', 'None']:
            unit_size = '-'

        rows_text += f"{i}. {sku} | {name} | Qty: {qty} | ${price:.2f} | Subtotal: ${subtotal:.2f}\n"
        if img:
            rows_text += f"   Image: {img}\n"

        rows_html += f"""<tr>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;font-size:13px;">{i}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;font-size:13px;font-weight:bold;">{sku}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;font-size:13px;">{name}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;font-size:13px;">{qty}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;font-size:13px;">${price:.2f}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:right;font-size:13px;font-weight:bold;">${subtotal:.2f}</td>
            <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;font-size:13px;">{f'<a href="{img}" style="color:#0563C1;">View</a>' if img else '-'}</td>
        </tr>"""

    subject = f"【Inquiry Received】Thank you for your inquiry - {pi_no}"

    text_body = f"""Dear {contact.get('name', 'Valued Customer')},

Thank you for your inquiry! We have received your request and will get back to you shortly.

Your Inquiry Reference: {pi_no}

PRODUCTS INQUIRY ({len(cart)} items)
----------------------------------------
{rows_text}
----------------------------------------
TOTAL: ${total:.2f}

Our team will prepare a formal PI (Proforma Invoice) and send it to you via email.

Best regards,
PARTY MAKER
info@partymaker.cn
"""

    html_body = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;max-width:720px;margin:0 auto;padding:0;background:#f5f5f5;">
    <div style="background:#9CAF88;color:white;padding:28px 32px;border-radius:12px 12px 0 0;">
        <h1 style="margin:0;font-size:22px;">✅ Inquiry Received</h1>
        <p style="margin:8px 0 0;opacity:0.9;font-size:14px;">Reference: {pi_no} | {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
    </div>
    <div style="background:white;padding:24px 32px;border:1px solid #ddd;border-top:none;">
        <p style="font-size:15px;color:#333;">Dear <strong>{contact.get('name', 'Valued Customer')}</strong>,</p>
        <p style="font-size:14px;color:#555;line-height:1.6;">Thank you for your inquiry! We have received your request and will get back to you shortly with a formal <strong>Proforma Invoice (PI)</strong>.</p>
        <p style="font-size:14px;color:#555;line-height:1.6;">Below is a summary of your inquiry:</p>
    </div>
    <div style="background:white;padding:0 32px 24px;border:1px solid #ddd;border-top:none;">
        <table style="width:100%;border-collapse:collapse;font-size:13px;">
            <thead><tr style="background:#9CAF88;color:white;">
                <th style="padding:10px 12px;text-align:center;">#</th>
                <th style="padding:10px 12px;text-align:center;">SKU</th>
                <th style="padding:10px 12px;text-align:left;">Product</th>
                <th style="padding:10px 12px;text-align:center;">Qty</th>
                <th style="padding:10px 12px;text-align:right;">Price</th>
                <th style="padding:10px 12px;text-align:right;">Subtotal</th>
                <th style="padding:10px 12px;text-align:center;">Image</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
            <tfoot><tr style="background:#F7F9F5;">
                <td colspan="5"></td>
                <td style="padding:10px 12px;text-align:right;font-weight:bold;font-size:15px;">${total:.2f}</td>
                <td></td>
            </tr></tfoot>
        </table>
    </div>
    <div style="background:white;padding:20px 32px;border:1px solid #ddd;border-top:none;">
        <p style="font-size:14px;color:#555;line-height:1.6;">We will send you the formal PI via email within <strong>24 hours</strong>.</p>
        <p style="font-size:14px;color:#555;">If you have any questions, feel free to reply to this email.</p>
    </div>
    <div style="text-align:center;font-size:12px;color:#999;padding:16px;">
        Sent by <a href="https://party-maker-website.pages.dev" style="color:#9CAF88;">Party Maker</a>
    </div>
</body></html>"""

    # 直接用SMTP发送（CF Worker支持urllib HTTP请求，但SMTP需要socket）
    # 由于CF Worker无法直接SMTP，改为POST给VPS中转
    return notify_vps_simple({
        'action': 'customer_email',
        'pi_no': pi_no,
        'to_email': to_email,
        'subject': subject,
        'text_body': text_body,
        'html_body': html_body,
        'contact': contact,
        'cart': cart,
        'excel_b64': base64.b64encode(excel_bytes).decode('utf-8'),
        'send_attachment': False,
    })


def notify_vps_simple(payload):
    """POST给VPS发邮件"""
    import urllib.request
    import urllib.error

    req = urllib.request.Request(
        VPS_URL,
        data=json.dumps(payload).encode('utf-8'),
        headers={
            "Content-Type": "application/json",
            "X-Trigger-Key": TRIGGER_KEY
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read())
            return result
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')
        return {'sent': False, 'error': f'HTTP {e.code}: {body}'}
    except Exception as e:
        return {'sent': False, 'error': str(e)}


def notify_vps(payload):
    """POST Excel数据到VPS → VPS发info邮件（含附件含成本价）"""
    import urllib.request
    import urllib.error

    req = urllib.request.Request(
        VPS_URL,
        data=json.dumps(payload).encode('utf-8'),
        headers={
            "Content-Type": "application/json",
            "X-Trigger-Key": TRIGGER_KEY
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read())
            return result
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')
        return {'sent': False, 'error': f'HTTP {e.code}: {body}'}
    except Exception as e:
        return {'sent': False, 'error': str(e)}


def generate_excel(contact, cart, pi_no, now):
    """Generate Proforma Invoice Excel（含Cost/CNY列）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"

    sage_green = PatternFill(start_color="FF9CAF88", end_color="FF9CAF88", fill_type="solid")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    title_font = Font(name='Arial', size=16, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=11, bold=True)

    thin = Side(style='thin', color='FF000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 标题
    ws.merge_cells('A1:L1')
    ws['A1'] = 'PROFORMA INVOICE'
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # 发货人/收货人
    ws['A2'] = 'TO:'; ws['B2'] = contact.get('company', '')
    ws['G2'] = 'From:'; ws['H2'] = 'PARTY MAKER'
    ws['A3'] = 'ATTN:'; ws['B3'] = contact.get('name', '')
    ws['G3'] = 'ATTN:'; ws['H3'] = ''
    ws['A4'] = 'TEL:'; ws['B4'] = contact.get('phone', '')
    ws['G4'] = 'TEL:'; ws['H4'] = '+86-572-222222'
    ws['A5'] = 'EMAIL:'; ws['B5'] = contact.get('email', '')
    ws['G5'] = 'Email:'; ws['H5'] = 'info@partymaker.cn'
    ws['A6'] = 'COUNTRY:'; ws['B6'] = contact.get('country', '')
    ws['A7'] = 'REMARK:'; ws['B7'] = (contact.get('message', '') or '')[:100]
    ws['G7'] = 'PI No.'; ws['H7'] = pi_no

    for row in range(2, 8):
        for col in ['A', 'B', 'G', 'H']:
            ws[f'{col}{row}'].font = normal_font

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[7].height = 20

    # 表头（含Cost列）
    headers = [
        ('No.', 5), ('Item No.', 12), ('Image', 9.375), ('Product Name', 30),
        ('Description', 25), ('USD Price', 10), ('Qty', 8), ('Amount', 12),
        ('Cost(CNY)', 10), ('Unit Size', 12), ('Unit Weight', 10), ('Inner Size', 12),
    ]
    row = 9
    header_row = row
    for col_idx, (header, width) in enumerate(headers, 1):
        cl = get_cell_letter(col_idx)
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font; cell.fill = sage_green
        cell.alignment = center; cell.border = border
        ws.column_dimensions[cl].width = width
    ws.row_dimensions[row].height = 22
    row += 1

    # 产品数据
    total = 0
    for i, item in enumerate(cart, 1):
        qty = int(item.get('quantity', 0) or 0)
        price = float(item.get('price', 0) or 0)
        amount = qty * price
        total += amount
        sku = item.get('sku', item.get('id', '-'))
        name = item.get('name', '')
        desc = (item.get('description', '') or '').replace('\n', ' ')
        cost_price = float(item.get('_costPrice', 0) or 0)
        unit_size = item.get('_unitSize', '')
        if str(unit_size) in ['nan', None, '', 'None']:
            unit_size = ''

        images = item.get('images', [])
        img_url = (images[0] if images and isinstance(images, list) and images[0] else '')

        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=sku).alignment = center
        ws.cell(row=row, column=4, value=name).alignment = left
        ws.cell(row=row, column=5, value=desc).alignment = left
        c = ws.cell(row=row, column=6, value=price)
        c.number_format = '$#,##0.00'; c.alignment = right
        ws.cell(row=row, column=7, value=qty).alignment = center
        c = ws.cell(row=row, column=8, value=f'=F{row}*G{row}')
        c.number_format = '$#,##0.00'; c.alignment = right
        # Cost列：含成本价
        c = ws.cell(row=row, column=9, value=cost_price if cost_price > 0 else '')
        c.number_format = '¥#,##0.00'; c.alignment = right
        ws.cell(row=row, column=10, value=unit_size).alignment = center
        ws.cell(row=row, column=11, value='').alignment = center
        ws.cell(row=row, column=12, value='').alignment = center

        # 图片超链接
        if img_url:
            c = ws.cell(row=row, column=3)
            c.value = img_url; c.hyperlink = img_url
            c.font = Font(name='Arial', size=9, color='FF0563C1', underline='single')
            c.alignment = center
        else:
            ws.cell(row=row, column=3, value='-').alignment = center

        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font; cell.border = border
        ws.row_dimensions[row].height = 55
        row += 1

    # 总计行
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="TOTAL:").font = bold_font
    ws.cell(row=row, column=1).alignment = right
    c = ws.cell(row=row, column=8, value=f'=SUM(H{header_row+1}:H{row-1})')
    c.number_format = '$#,##0.00'; c.font = bold_font; c.alignment = right
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = border
    ws.row_dimensions[row].height = 22
    row += 2

    # Terms
    ws.cell(row=row, column=1, value="TERMS & CONDITIONS").font = bold_font
    row += 1
    for term in [
        "1. FOB Ningbo / Shanghai",
        "2. Price does not include testing, inspection and auditing costs",
        "3. Production time: 45 days after deposit is received",
        "4. Payment: 30% deposit, 70% balance before goods leave factory",
    ]:
        ws.cell(row=row, column=1, value=term).font = normal_font
        row += 1

    # Bank
    row += 1
    ws.cell(row=row, column=1, value="BANK INFORMATION").font = bold_font
    row += 1
    for label, val in [
        ("BENEFICIARY:", "JIATAO INDUSTRY (SHANGHAI) CO.,LTD"),
        ("BANK NAME:", "AGRICULTURAL BANK OF CHINA SHANGHAI YANGPU BRANCH"),
        ("BANK ADDRESS:", "NO.1128, XIANGYIN ROAD, YANGPU DISTRICT, SHANGHAI CHINA"),
        ("POST CODE:", "200433"),
        ("A/C NO.:", "09421014040006209"),
        ("SWIFT CODE:", "ABOCCNBJ090"),
    ]:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=val).font = normal_font
        row += 1

    # 签名
    row += 2
    ws.cell(row=row, column=1, value="BUYER SIGNATURE:").font = bold_font
    ws.cell(row=row, column=6, value="SELLER SIGNATURE:").font = bold_font
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = '_________________________'
    ws.merge_cells(f'F{row}:I{row}')
    ws[f'F{row}'] = '_________________________'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def json_response(data, status=200):
    from urllib.request import Response
    body = json.dumps(data).encode()
    headers = {
        "Content-Type": "application/json",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type"
    }
    return Response(body, status=status, headers=headers)


def cors_response():
    from urllib.request import Response
    headers = {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type"
    }
    return Response(b'', status=204, headers=headers)
