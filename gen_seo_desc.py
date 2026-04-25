"""
Generate SEO descriptions for all products using AI-powered local templates.
Creates rich, varied descriptions from product name + description + theme + subcategory.
"""

import json
import os
import re
import random
import hashlib

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WEBSITE_DIR = BASE_DIR
PRODUCTS_JSON = os.path.join(BASE_DIR, 'products.json')
OUTPUT_XLSX = os.path.join(BASE_DIR, '..', 'seo_desc_review.xlsx')

def load_products():
    with open(PRODUCTS_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data['products']

def clean_text(text):
    text = text.strip()
    text = re.sub(r'\n+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text

def pick_by_hash(sku, options):
    """Deterministically pick an option based on SKU hash."""
    idx = int(hashlib.md5(str(sku).encode()).hexdigest(), 16) % len(options)
    return options[idx]

def generate_seo_desc(products):
    """
    Generate varied, natural-sounding SEO descriptions.
    Uses many template variations + field-aware assembly to avoid repetition.
    """
    
    # Theme intros - how to frame the product in context
    theme_intros = {
        'Ramadan': [
            "Celebrate the holy month with this {name}",
            "Bring Ramadan spirit to your home with this {name}",
            "Enhance your Eid Mubarak festivities with this {name}",
            "This beautiful {name} is designed for Ramadan and Eid celebrations",
            "Create a warm, festive atmosphere during Ramadan with this {name}",
            "A stunning {name} to decorate your space for the blessed month",
            "Make this Ramadan memorable with this {name}",
            "Adorn your home with this {name} for the holy month of Ramadan",
        ],
        'Christmas': [
            "Add holiday magic with this {name}",
            "This {name} brings Christmas cheer to any space",
            "Transform your home for the holidays with this {name}",
            "Create a winter wonderland with this festive {name}",
            "A must-have {name} for your Christmas celebration",
        ],
        'Halloween': [
            "Spook up your party with this {name}",
            "This {name} sets the perfect Halloween mood",
            "Create an eerie atmosphere with this {name}",
            "A thrilling {name} for your Halloween decorations",
        ],
        'New Year': [
            "Ring in the new year with this {name}",
            "This {name} adds sparkle to your New Year celebration",
            "Start the year right with this festive {name}",
            "A dazzling {name} for your New Year Eve party",
        ],
        'Birthday': [
            "Make every birthday special with this {name}",
            "This {name} is perfect for birthday celebrations",
            "Add fun and color to birthday parties with this {name}",
            "A cheerful {name} designed for unforgettable birthdays",
        ],
        'Wedding': [
            "Add elegance to your wedding with this {name}",
            "This {name} creates a romantic atmosphere for your special day",
            "A sophisticated {name} perfect for wedding receptions",
            "Make your wedding day unforgettable with this {name}",
        ],
        'Valentine': [
            "Express your love with this {name}",
            "This {name} sets a romantic mood for Valentine's Day",
            "A heartfelt {name} perfect for Valentine celebrations",
            "Surprise your loved one with this {name}",
        ],
        'Baby Shower': [
            "Welcome the little one with this adorable {name}",
            "This {name} adds sweetness to baby shower celebrations",
            "A charming {name} for your baby shower party",
            "Celebrate new arrivals with this delightful {name}",
        ],
    }
    
    # Use-case closers
    use_closers = [
        "Ideal for homes, event venues, and party spaces.",
        "Great for both indoor and outdoor celebrations.",
        "A wonderful choice for party planners and event decorators.",
        "Perfect for wholesale buyers looking for quality party supplies.",
        "Suitable for retail stores and bulk party supply orders.",
        "Bulk available for event planners and party supply distributors.",
        "A popular choice among party supply wholesalers worldwide.",
        "Available for bulk purchase with competitive pricing.",
        "An excellent addition to any party supply catalog.",
        "Perfect for creating memorable celebration experiences.",
        "Great value for party supply retailers and event organizers.",
        "Designed to meet wholesale party supply needs.",
    ]
    
    # Subcategory feature phrases
    subcat_features = {
        'LED Light': 'It features energy-efficient LED illumination that creates a warm, inviting glow.',
        'Lantern': 'This classic lantern design adds an elegant, traditional touch to any setting.',
        'Banner': 'An eye-catching banner that instantly transforms walls and doorways.',
        'Backdrop': 'Use it as a stunning photo backdrop to create memorable celebration moments.',
        'Garland': 'This beautiful garland can be draped across tables, walls, or mantels.',
        'Cake Topper': 'Simply place on cakes and cupcakes for an instant themed decoration.',
        'Candle': 'It creates warm ambient candlelight perfect for intimate gatherings.',
        'Balloon': 'This balloon design adds vibrant color and festive energy to any space.',
        'Tablecloth': 'A premium table covering that sets the perfect party mood in seconds.',
        'Cup': 'Serve beverages in style with these themed festive cups.',
        'Plate': 'Coordinated themed tableware that completes your party table setting.',
        'Napkin': 'These coordinated table accessories add a polished finishing touch.',
        'Pinata': 'A fun interactive party activity that guests of all ages will enjoy.',
        'Deco-Hanging': 'An elegant hanging ornament that adds charm to any space.',
        'Deco-Wood': 'Crafted from natural wood, it brings a warm, organic feel to your decor.',
        'Sticker': 'These decorative stickers are perfect for gifts, cards, and party favors.',
        'Confetti': 'Scatter for a celebratory finishing touch on tables and surfaces.',
        'Gift Bag': 'Beautiful gift packaging that makes every present feel extra special.',
        'Invitation': 'Themed party invitations that set the tone for your upcoming celebration.',
        'Hat': 'Fun party headwear that guests will love wearing.',
        'Blower': 'Celebration party blowers that add excitement to any festivity.',
        'Flag': 'A vibrant flag decoration perfect for both indoor and outdoor display.',
        'Greeting Card': 'Themed greeting cards to share warm wishes with family and friends.',
        'Light': 'A decorative lighting piece that enhances the ambiance of any room.',
        'Picks': 'Decorative food picks and toppers that make treats look irresistible.',
        'Tag': 'Themed gift tags and labels that add a personal touch to presents.',
        'Wrapping': 'Festive gift wrapping that makes every gift stand out.',
        'Crown': 'A royal party headpiece perfect for the guest of honor.',
        'Sash': 'A celebratory sash accessory that makes any occasion feel special.',
        'Photo Frame': 'A party photo booth prop that guarantees fun photos.',
        'Cutout': 'A standup decoration display that adds character to any venue.',
    }
    
    # Material hints based on description keywords
    material_patterns = [
        (r'polyester', 'Made from premium polyester fabric'),
        (r'paper', 'Crafted from quality paper materials'),
        (r'wood', 'Made from natural wood'),
        (r'acrylic', 'Made from durable acrylic'),
        (r'metal', 'Crafted from quality metal'),
        (r'plastic', 'Made from safe plastic materials'),
        (r'foam', 'Made from lightweight foam'),
        (r'cloth|fabric', 'Made from quality fabric'),
        (r'pvc', 'Made from PVC material'),
        (r'glass', 'Made from glass'),
        (r'ceramic', 'Crafted from ceramic'),
        (r'silk|satin', 'Made from luxurious silk/satin'),
        (r'latex', 'Made from natural latex'),
        (r'cardboard|card stock', 'Made from sturdy cardboard'),
        (r'felt', 'Made from soft felt material'),
        (r'flannel|plush', 'Made from soft plush fabric'),
        (r'cotton', 'Made from cotton'),
        (r'ribbon', 'Featuring decorative ribbon details'),
        (r'glitter', 'with glitter accents'),
        (r'led|LED', 'with LED lighting'),
        (r'solar', 'with solar-powered functionality'),
        (r'battery', 'battery-operated for convenience'),
    ]
    
    # Function phrases based on subcategory
    function_phrases = {
        'Lantern': 'Hang it to create an inviting glow in any room.',
        'LED Light': 'Illuminate your space with a warm, festive glow.',
        'Banner': 'Display it across walls or doorways for instant festive decor.',
        'Backdrop': 'Set up a beautiful photo area for guests to capture memories.',
        'Garland': 'Drape it across tables, walls, or mantels for a decorative touch.',
        'Cake Topper': 'Place it on cakes and cupcakes for an instant themed look.',
        'Tablecloth': 'Cover your dining table to set the perfect party mood.',
        'Balloon': 'Inflate and arrange for a colorful party atmosphere.',
        'Pinata': 'Fill with treats and candies for a fun party game.',
        'Flag': 'Display indoors or outdoors to show your festive spirit.',
        'Sticker': 'Peel and stick on gifts, cards, or party favors.',
        'Confetti': 'Scatter on tables for a celebratory finishing touch.',
        'Gift Bag': 'Fill with party favors and gifts for your guests.',
        'Deco-Hanging': 'Hang from ceilings, doorways, or walls.',
        'Deco-Wood': 'Place on tables or shelves as a decorative centerpiece.',
        'Cup': 'Serve drinks in style with these themed cups.',
        'Plate': 'Serve snacks and meals on these themed plates.',
        'Napkin': 'Complete your table setting with coordinated napkins.',
        'Crown': 'Wear it to feel like royalty at your celebration.',
        'Cutout': 'Stand it up as a photo prop or room decoration.',
    }
    
    results = []
    total = len(products)
    
    for i, p in enumerate(products):
        sku = p.get('sku') or p.get('id', '')
        name = p.get('name', '').strip()
        desc = clean_text(p.get('description', ''))
        theme = p.get('theme', 'General Party')
        subcat = p.get('subcategory', '')
        
        if not name:
            results.append({'sku': sku, 'name': name, 'description': desc, 'theme': theme, 'subcategory': subcat, 'seo_desc': ''})
            continue
        
        sentences = []
        
        # === Sentence 1: Theme-intro + specs ===
        intros = theme_intros.get(theme, theme_intros.get('Birthday', [
            "Enhance your celebration with this {name}",
            "This {name} is perfect for parties and events",
            "Add a festive touch with this {name}",
            "A great {name} for any celebration or event",
        ]))
        intro = pick_by_hash(sku, intros).format(name=name)
        
        # Add specs if available
        if desc and len(desc) > 2:
            # Try to extract size info
            size_match = re.search(r'(\d+\s*[xX*]\s*\d+(?:\s*[xX*]\s*\d+)?\s*cm)', desc)
            qty_match = re.search(r'(\d+)\s*pcs', desc, re.IGNORECASE)
            
            specs_parts = []
            if size_match:
                specs_parts.append(size_match.group(1))
            if qty_match and 'ctn' not in desc.lower():
                specs_parts.append(f"{qty_match.group(1)} pcs")
            
            if specs_parts:
                spec_str = ' (' + ', '.join(specs_parts) + ')'
                sentences.append(intro + spec_str + '.')
            else:
                # Just append key detail
                short_desc = desc[:60]
                if len(desc) > 60:
                    short_desc = short_desc[:57] + '...'
                sentences.append(f"{intro}, {short_desc}.")
        else:
            sentences.append(intro + '.')
        
        # === Sentence 2: Subcategory feature OR Material ===
        feature_sentence = ''
        for key, phrase in subcat_features.items():
            if key.lower() in subcat.lower():
                feature_sentence = phrase
                break
        
        if not feature_sentence:
            # Fallback: try to detect material from description
            for pattern, phrase in material_patterns:
                if re.search(pattern, desc, re.IGNORECASE):
                    feature_sentence = phrase + '.'
                    break
        
        if not feature_sentence:
            feature_sentence = pick_by_hash(sku, [
                "Crafted with attention to detail and quality materials.",
                "Made with quality materials for lasting enjoyment.",
                "Designed for both visual appeal and practical use.",
                "A well-crafted party supply built to impress your guests.",
            ])
        
        sentences.append(feature_sentence)
        
        # === Sentence 3: Use case / function (skip if subcat feature already covers it) ===
        has_subcat_feature = any(key.lower() in subcat.lower() for key in subcat_features)
        if has_subcat_feature:
            # Subcat feature was used in sentence 2, add a general closer instead
            closer = pick_by_hash(sku, use_closers)
            sentences.append(closer)
        else:
            func = None
            for key, phrase in function_phrases.items():
                if key.lower() in subcat.lower():
                    func = phrase
                    break
            if func:
                sentences.append(func)
            else:
                closer = pick_by_hash(sku, use_closers)
                sentences.append(closer)
        
        seo_desc = ' '.join(sentences)
        # Clean up
        seo_desc = re.sub(r'\.\.', '.', seo_desc)
        seo_desc = re.sub(r'\s+', ' ', seo_desc).strip()
        # Capitalize first letter
        if seo_desc and seo_desc[0].islower():
            seo_desc = seo_desc[0].upper() + seo_desc[1:]
        
        results.append({
            'sku': sku,
            'name': name,
            'description': desc,
            'theme': theme,
            'subcategory': subcat,
            'seo_desc': seo_desc,
        })
        
        if (i + 1) % 200 == 0:
            print(f"  Progress: {i+1}/{total}")
    
    return results

def save_to_excel(results):
    """Save results to Excel for review."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SEO Descriptions'
    
    headers = ['SKU', 'Product Name', 'Original Description', 'Theme', 'Subcategory', 'Generated SEO Description']
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4A5A3A', end_color='4A5A3A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for row_idx, item in enumerate(results, 2):
        values = [
            item['sku'],
            item['name'],
            item.get('description', ''),
            item.get('theme', ''),
            item.get('subcategory', ''),
            item.get('seo_desc', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = wrap_alignment
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 75
    
    ws.freeze_panes = 'A2'
    
    wb.save(OUTPUT_XLSX)
    print(f"\nSaved to: {OUTPUT_XLSX}")

def main():
    print("Loading products...")
    products = load_products()
    print(f"Total products: {len(products)}")
    
    print("\nGenerating SEO descriptions...")
    results = generate_seo_desc(products)
    
    print("\nSaving to Excel for review...")
    save_to_excel(results)
    
    # Show examples
    print("\n=== Sample SEO Descriptions ===")
    random.seed(42)
    samples = random.sample(results, min(8, len(results)))
    for s in samples:
        print(f"\nSKU {s['sku']}: {s['name']}")
        print(f"  Theme: {s['theme']} / {s['subcategory']}")
        print(f"  Original: {s['description'][:60]}...")
        print(f"  SEO: {s['seo_desc']}")

if __name__ == '__main__':
    main()
